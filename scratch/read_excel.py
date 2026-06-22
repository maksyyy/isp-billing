import openpyxl
import os

def main():
    excel_path = os.path.abspath("Analisis biaya dan manfaat.xlsx")
    output_path = os.path.abspath("Analisis biaya dan manfaat.md")
    
    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"# Analisis Biaya dan Manfaat\n\n")
        f.write(f"File: `{excel_path}`\n\n")
        f.write(f"Sheets: {', '.join(wb.sheetnames)}\n\n")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            f.write(f"## Sheet: {sheet_name}\n\n")
            
            # Read all cells
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                f.write("*(Empty sheet)*\n\n")
                continue
                
            # Determine the maximum length row to normalize the table columns
            max_cols = max(len(row) for row in rows)
            
            # First, find actual data boundaries (ignore trailing completely None rows)
            last_non_empty_row = -1
            for idx, row in enumerate(rows):
                if any(cell is not None for cell in row):
                    last_non_empty_row = idx
                    
            if last_non_empty_row == -1:
                f.write("*(Empty sheet)*\n\n")
                continue
                
            valid_rows = rows[:last_non_empty_row + 1]
            
            # Format rows
            markdown_rows = []
            for row in valid_rows:
                formatted_row = []
                for cell in row:
                    if cell is None:
                        formatted_row.append("")
                    else:
                        # format floats nicely
                        if isinstance(cell, float):
                            if cell.is_integer():
                                formatted_row.append(str(int(cell)))
                            else:
                                formatted_row.append(f"{cell:.2f}")
                        else:
                            formatted_row.append(str(cell).replace("\n", "<br>").replace("|", "\\|"))
                # Pad row if needed
                while len(formatted_row) < max_cols:
                    formatted_row.append("")
                markdown_rows.append(formatted_row)

            if not markdown_rows:
                f.write("*(Empty sheet)*\n\n")
                continue

            # Print table
            # Use first row as header if it's there
            headers = markdown_rows[0]
            f.write("| " + " | ".join(headers) + " |\n")
            f.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
            for row in markdown_rows[1:]:
                f.write("| " + " | ".join(row) + " |\n")
            f.write("\n\n")
            
    print(f"Successfully wrote to {output_path}")

if __name__ == "__main__":
    main()
