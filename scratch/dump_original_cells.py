import openpyxl
import os

def main():
    excel_path = os.path.abspath("Analisis biaya dan manfaat (Contoh).xlsx")
    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=False) # Keep formulas!
    sheet = wb.active
    print(f"Sheet name: {sheet.title}")
    
    # Iterate through all rows and columns up to max row and column
    for row_idx in range(1, sheet.max_row + 1):
        row_vals = []
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                row_vals.append((col_idx, cell.value))
        if row_vals:
            print(f"Row {row_idx}: {row_vals}")

if __name__ == "__main__":
    main()
