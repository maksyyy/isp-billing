import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    workspace_dir = os.path.abspath(".")
    new_excel_path = os.path.join(workspace_dir, "Analisis biaya dan manfaat.xlsx")
    markdown_path = os.path.join(workspace_dir, "Analisis Biaya dan Manfaat Proyek.md")

    # Recreate the workbook from scratch to extend up to 5 years beautifully
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis Biaya dan manfaat"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions (Premium theme matching the original template style)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_italic = Font(name="Calibri", size=10, italic=True)
    font_title = Font(name="Calibri", size=14, bold=True, color="1B365D")

    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_section = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Title
    ws["A1"] = "Analisis Biaya dan Manfaat Proyek ISP Billing"
    ws["A1"].font = font_title
    ws.row_dimensions[1].height = 25

    # Headers
    headers = ["Keterangan", "Tahun 0", "Tahun 1", "Tahun 2", "Tahun 3", "Tahun 4", "Tahun 5"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = font_bold
        cell.alignment = align_center if col_idx > 1 else align_left
        cell.border = thin_border
    ws.row_dimensions[3].height = 22

    # Data Rows
    # Structure: (Row, Col, Value, Format)
    # Row 2: BIAYA-BIAYA
    ws.cell(row=4, column=1, value="BIAYA-BIAYA").font = font_bold
    
    # Row 3: Biaya pengembangan sistem
    ws.cell(row=5, column=1, value="Biaya pengembangan sistem").font = font_bold
    
    # Row 4: Biaya pengadaan
    ws.cell(row=6, column=1, value="Biaya pengadaan")
    ws.cell(row=6, column=2, value=40000000.0)
    ws.cell(row=6, column=3, value=0.0)
    ws.cell(row=6, column=4, value=0.0)
    ws.cell(row=6, column=5, value=0.0)
    ws.cell(row=6, column=6, value=0.0)
    ws.cell(row=6, column=7, value=0.0)

    # Row 5: Biaya persiapan operasi
    ws.cell(row=7, column=1, value="Biaya persiapan operasi")
    ws.cell(row=7, column=2, value=4000000.0)
    ws.cell(row=7, column=3, value=0.0)
    ws.cell(row=7, column=4, value=0.0)
    ws.cell(row=7, column=5, value=0.0)
    ws.cell(row=7, column=6, value=0.0)
    ws.cell(row=7, column=7, value=0.0)

    # Row 6: Biaya proyek
    ws.cell(row=8, column=1, value="Biaya proyek").font = font_bold

    # Row 7: Biaya konsultan
    ws.cell(row=9, column=1, value="Biaya konsultan")
    ws.cell(row=9, column=2, value=10000000.0)
    ws.cell(row=9, column=3, value=0.0)
    ws.cell(row=9, column=4, value=0.0)
    ws.cell(row=9, column=5, value=0.0)
    ws.cell(row=9, column=6, value=0.0)
    ws.cell(row=9, column=7, value=0.0)

    # Row 8: Tahap analisis sistem
    ws.cell(row=10, column=1, value="Tahap analisis sistem")
    ws.cell(row=10, column=2, value=8000000.0)
    ws.cell(row=10, column=3, value=0.0)
    ws.cell(row=10, column=4, value=0.0)
    ws.cell(row=10, column=5, value=0.0)
    ws.cell(row=10, column=6, value=0.0)
    ws.cell(row=10, column=7, value=0.0)

    # Row 9: Tahap desain sistem
    ws.cell(row=11, column=1, value="Tahap desain sistem")
    ws.cell(row=11, column=2, value=8000000.0)
    ws.cell(row=11, column=3, value=0.0)
    ws.cell(row=11, column=4, value=0.0)
    ws.cell(row=11, column=5, value=0.0)
    ws.cell(row=11, column=6, value=0.0)
    ws.cell(row=11, column=7, value=0.0)

    # Row 10: Penerapan sistem
    ws.cell(row=12, column=1, value="Penerapan sistem  ")
    ws.cell(row=12, column=2, value=15000000.0)
    ws.cell(row=12, column=3, value=0.0)
    ws.cell(row=12, column=4, value=0.0)
    ws.cell(row=12, column=5, value=0.0)
    ws.cell(row=12, column=6, value=0.0)
    ws.cell(row=12, column=7, value=0.0)

    # Row 11: Biaya penyusutan
    ws.cell(row=13, column=1, value="Biaya penyusutan")
    ws.cell(row=13, column=2, value=0.0)
    ws.cell(row=13, column=3, value=0.0)
    ws.cell(row=13, column=4, value=0.0)
    ws.cell(row=13, column=5, value=0.0)
    ws.cell(row=13, column=6, value=0.0)
    ws.cell(row=13, column=7, value=300000.0) # penyusutan in Year 5

    # Row 12: Total biaya proyek
    ws.cell(row=14, column=1, value="Total biaya proyek  ").font = font_bold
    ws.cell(row=14, column=2, value="=sum(B9:B12)")
    ws.cell(row=14, column=3, value=0.0)
    ws.cell(row=14, column=4, value=0.0)
    ws.cell(row=14, column=5, value=0.0)
    ws.cell(row=14, column=6, value=0.0)
    ws.cell(row=14, column=7, value="=G13")

    # Row 14: Total Biaya Pengembangan Sistem
    ws.cell(row=16, column=1, value="Total Biaya Pengembangan Sistem").font = font_bold
    ws.cell(row=16, column=2, value="=sum(B6:B12)")
    ws.cell(row=16, column=3, value=0.0)
    ws.cell(row=16, column=4, value=0.0)
    ws.cell(row=16, column=5, value=0.0)
    ws.cell(row=16, column=6, value=0.0)
    ws.cell(row=16, column=7, value="=G14")

    # Row 16: Biaya Operasional dan Perawatan
    ws.cell(row=18, column=1, value="Biaya Operasional dan Perawatan").font = font_bold
    
    # Row 17: Operasional
    ws.cell(row=19, column=1, value="Operasional")
    ws.cell(row=19, column=2, value=0.0)
    ws.cell(row=19, column=3, value=3000000.0)
    ws.cell(row=19, column=4, value=3500000.0)
    ws.cell(row=19, column=5, value=4000000.0)
    ws.cell(row=19, column=6, value=4500000.0)
    ws.cell(row=19, column=7, value=5000000.0)

    # Row 18: Perawatan
    ws.cell(row=20, column=1, value="Perawatan")
    ws.cell(row=20, column=2, value=0.0)
    ws.cell(row=20, column=3, value=2000000.0)
    ws.cell(row=20, column=4, value=2500000.0)
    ws.cell(row=20, column=5, value=2500000.0)
    ws.cell(row=20, column=6, value=3000000.0)
    ws.cell(row=20, column=7, value=3000000.0)

    # Row 20: Total Biaya Operasional dan Perawatan
    ws.cell(row=22, column=1, value="Total Biaya Operasional dan Perawatan").font = font_bold
    ws.cell(row=22, column=2, value=0.0)
    ws.cell(row=22, column=3, value="=sum(C19:C20)")
    ws.cell(row=22, column=4, value="=sum(D19:D20)")
    ws.cell(row=22, column=5, value="=sum(E19:E20)")
    ws.cell(row=22, column=6, value="=sum(F19:F20)")
    ws.cell(row=22, column=7, value="=sum(G19:G20)")

    # Row 22: TOTAL BIAYA
    ws.cell(row=24, column=1, value="TOTAL BIAYA").font = font_bold
    ws.cell(row=24, column=2, value="=B16")
    ws.cell(row=24, column=3, value="=C22")
    ws.cell(row=24, column=4, value="=D22")
    ws.cell(row=24, column=5, value="=E22")
    ws.cell(row=24, column=6, value="=F22")
    ws.cell(row=24, column=7, value="=G22+G16") # Year 5 has operasional + penyusutan

    # Row 24: MANFAAT
    ws.cell(row=26, column=1, value="MANFAAT").font = font_bold
    
    # Row 25: Berwujud
    ws.cell(row=27, column=1, value="Berwujud").font = font_bold
    
    # Row 26: Penghematan biaya operasional pershn
    ws.cell(row=28, column=1, value="Penghematan biaya operasional pershn")
    ws.cell(row=28, column=2, value=0.0)
    ws.cell(row=28, column=3, value=10000000.0)
    ws.cell(row=28, column=4, value=12000000.0)
    ws.cell(row=28, column=5, value=15000000.0)
    ws.cell(row=28, column=6, value=18000000.0)
    ws.cell(row=28, column=7, value=22000000.0)

    # Row 27: Peningkatan penjualan
    ws.cell(row=29, column=1, value="Peningkatan penjualan")
    ws.cell(row=29, column=2, value=0.0)
    ws.cell(row=29, column=3, value=12000000.0)
    ws.cell(row=29, column=4, value=20000000.0)
    ws.cell(row=29, column=5, value=30000000.0)
    ws.cell(row=29, column=6, value=45000000.0)
    ws.cell(row=29, column=7, value=60000000.0)

    # Row 28: Penurunan biaya persediaan
    ws.cell(row=30, column=1, value="Penurunan biaya persediaan")
    ws.cell(row=30, column=2, value=0.0)
    ws.cell(row=30, column=3, value=4000000.0)
    ws.cell(row=30, column=4, value=5000000.0)
    ws.cell(row=30, column=5, value=6000000.0)
    ws.cell(row=30, column=6, value=7000000.0)
    ws.cell(row=30, column=7, value=8000000.0)

    # Row 29: Tak Berwujud
    ws.cell(row=31, column=1, value="Tak Berwujud").font = font_bold
    
    # Row 30: Peningkatan pelayanan
    ws.cell(row=32, column=1, value="Peningkatan pelayanan")
    ws.cell(row=32, column=2, value=0.0)
    ws.cell(row=32, column=3, value=5000000.0)
    ws.cell(row=32, column=4, value=7000000.0)
    ws.cell(row=32, column=5, value=10000000.0)
    ws.cell(row=32, column=6, value=12000000.0)
    ws.cell(row=32, column=7, value=15000000.0)

    # Row 31: Peningkatan kepuasan pekerjaan
    ws.cell(row=33, column=1, value="Peningkatan kepuasan pekerjaan")
    ws.cell(row=33, column=2, value=0.0)
    ws.cell(row=33, column=3, value=3000000.0)
    ws.cell(row=33, column=4, value=4000000.0)
    ws.cell(row=33, column=5, value=5000000.0)
    ws.cell(row=33, column=6, value=7000000.0)
    ws.cell(row=33, column=7, value=9000000.0)

    # Row 32: Peningkatan pengambilan keputusan
    ws.cell(row=34, column=1, value="Peningkatan pengambilan keputusan")
    ws.cell(row=34, column=2, value=0.0)
    ws.cell(row=34, column=3, value=4000000.0)
    ws.cell(row=34, column=4, value=5000000.0)
    ws.cell(row=34, column=5, value=6000000.0)
    ws.cell(row=34, column=6, value=8000000.0)
    ws.cell(row=34, column=7, value=10000000.0)

    # Row 34: TOTAL MANFAAT
    ws.cell(row=36, column=1, value="TOTAL MANFAAT").font = font_bold
    ws.cell(row=36, column=2, value=0.0)
    ws.cell(row=36, column=3, value="=SUM(C28:C34)")
    ws.cell(row=36, column=4, value="=SUM(D28:D34)")
    ws.cell(row=36, column=5, value="=SUM(E28:E34)")
    ws.cell(row=36, column=6, value="=SUM(F28:F34)")
    ws.cell(row=36, column=7, value="=SUM(G28:G34)")

    # Row 36: Proceed ( Selisih TB dan TM )
    ws.cell(row=38, column=1, value="Proceed ( Selisih TB dan TM )").font = font_bold
    ws.cell(row=38, column=2, value=0.0)
    ws.cell(row=38, column=3, value="=C36-C24")
    ws.cell(row=38, column=4, value="=D36-D24")
    ws.cell(row=38, column=5, value="=E36-E24")
    ws.cell(row=38, column=6, value="=F36-F24")
    ws.cell(row=38, column=7, value="=G36-G24")

    # Format numbers for general table
    for r in range(4, 39):
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and not isinstance(cell.value, str):
                cell.number_format = '#,##0'
            cell.border = thin_border
            if c > 1:
                cell.alignment = align_right

    # Row 41: Summary Table Header
    ws.cell(row=41, column=1, value="").font = font_bold
    ws.cell(row=41, column=2, value="Tahun 0").font = font_bold
    ws.cell(row=41, column=3, value="Tahun 1").font = font_bold
    ws.cell(row=41, column=4, value="Tahun 2").font = font_bold
    ws.cell(row=41, column=5, value="Tahun 3").font = font_bold
    ws.cell(row=41, column=6, value="Tahun 4").font = font_bold
    ws.cell(row=41, column=7, value="Tahun 5").font = font_bold
    ws.cell(row=41, column=8, value="Total ").font = font_bold
    ws.row_dimensions[41].height = 20

    # Row 42: TOTAL BIAYA Summary
    ws.cell(row=42, column=1, value="TOTAL BIAYA").font = font_bold
    ws.cell(row=42, column=2, value="=B24")
    ws.cell(row=42, column=3, value="=C24")
    ws.cell(row=42, column=4, value="=D24")
    ws.cell(row=42, column=5, value="=E24")
    ws.cell(row=42, column=6, value="=F24")
    ws.cell(row=42, column=7, value="=G24")
    ws.cell(row=42, column=8, value="=SUM(B42:G42)")

    # Row 43: TOTAL MANFAAT Summary
    ws.cell(row=43, column=1, value="TOTAL MANFAAT").font = font_bold
    ws.cell(row=43, column=2, value="=B36")
    ws.cell(row=43, column=3, value="=C36")
    ws.cell(row=43, column=4, value="=D36")
    ws.cell(row=43, column=5, value="=E36")
    ws.cell(row=43, column=6, value="=F36")
    ws.cell(row=43, column=7, value="=G36")
    ws.cell(row=43, column=8, value="=SUM(C43:G43)")

    # Row 44: Proceed Summary
    ws.cell(row=44, column=1, value="Proceed ( Selisih TB dan TM )").font = font_bold
    ws.cell(row=44, column=2, value=0)
    ws.cell(row=44, column=3, value="=C38")
    ws.cell(row=44, column=4, value="=D38")
    ws.cell(row=44, column=5, value="=E38")
    ws.cell(row=44, column=6, value="=F38")
    ws.cell(row=44, column=7, value="=G38")
    ws.cell(row=44, column=8, value="")

    for r in range(41, 45):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c > 1:
                cell.alignment = align_right
                if cell.value is not None and not isinstance(cell.value, str):
                    cell.number_format = '#,##0'

    # Payback Period Section (Row 47-59)
    ws.cell(row=47, column=1, value="Payback Period").font = font_bold
    
    ws.cell(row=50, column=1, value="Nilai Investasi")
    ws.cell(row=50, column=3, value="=B42") # Col C (3)

    ws.cell(row=51, column=1, value="Proceed tahun I")
    ws.cell(row=51, column=3, value="=C44") # Col C (3)

    ws.cell(row=52, column=1, value="Sisa investasi tahun II")
    ws.cell(row=52, column=3, value="=C50-C51") # Col C (3)

    ws.cell(row=53, column=1, value="Proceed tahun II")
    ws.cell(row=53, column=3, value="=D44") # Col C (3)

    ws.cell(row=54, column=1, value="Sisa investasi tahun III")
    ws.cell(row=54, column=3, value="=C52-C53") # Col C (3)

    ws.cell(row=57, column=1, value="Sisa")
    ws.cell(row=57, column=3, value="=C58")
    ws.cell(row=57, column=4, value="hari")
    ws.cell(row=57, column=5, value="=365*C57")

    ws.cell(row=58, column=3, value="=C54/E44") # sisa fraction C58

    ws.cell(row=59, column=1, value="Sisa investasi tahun ke 3 tertutup oleh proceed tahun ke 3 yang berarti investasi akan kembali pada tahun ke 3 yaitu tepatnya 2.08 tahun atau 2 tahun 28 hari")
    ws.cell(row=59, column=1).font = font_italic

    # Number format for Payback Period values
    for r in [50, 51, 52, 53, 54]:
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=3).alignment = align_right
        ws.cell(row=r, column=3).font = font_regular
        ws.cell(row=r, column=1).font = font_regular

    ws.cell(row=58, column=3).number_format = '0.00'
    ws.cell(row=58, column=3).alignment = align_right
    ws.cell(row=57, column=5).number_format = '0.00'
    ws.cell(row=57, column=5).alignment = align_right

    # ROI Section (Row 62-66)
    ws.cell(row=62, column=1, value="ROI").font = font_bold
    
    # Header
    ws.cell(row=63, column=2, value="Tahun 0").font = font_bold
    ws.cell(row=63, column=3, value="Tahun 1").font = font_bold
    ws.cell(row=63, column=4, value="Tahun 2").font = font_bold
    ws.cell(row=63, column=5, value="Tahun 3").font = font_bold
    ws.cell(row=63, column=6, value="Tahun 4").font = font_bold
    ws.cell(row=63, column=7, value="Tahun 5").font = font_bold
    ws.cell(row=63, column=8, value="Total ").font = font_bold
    ws.cell(row=63, column=9, value="selisih TB-TM").font = font_bold

    ws.cell(row=64, column=1, value="TOTAL BIAYA").font = font_bold
    ws.cell(row=64, column=2, value="=B42")
    ws.cell(row=64, column=3, value="=C42")
    ws.cell(row=64, column=4, value="=D42")
    ws.cell(row=64, column=5, value="=E42")
    ws.cell(row=64, column=6, value="=F42")
    ws.cell(row=64, column=7, value="=G42")
    ws.cell(row=64, column=8, value="=SUM(B64:G64)")
    ws.cell(row=64, column=9, value="=H65-H64")

    ws.cell(row=65, column=1, value="TOTAL MANFAAT").font = font_bold
    ws.cell(row=65, column=2, value="=B43")
    ws.cell(row=65, column=3, value="=C43")
    ws.cell(row=65, column=4, value="=D43")
    ws.cell(row=65, column=5, value="=E43")
    ws.cell(row=65, column=6, value="=F43")
    ws.cell(row=65, column=7, value="=G43")
    ws.cell(row=65, column=8, value="=SUM(C65:G65)")

    ws.cell(row=66, column=1, value="Proceed ( Selisih TB dan TM )").font = font_bold
    ws.cell(row=66, column=2, value=0)
    ws.cell(row=66, column=3, value="=C44")
    ws.cell(row=66, column=4, value="=D44")
    ws.cell(row=66, column=5, value="=E44")
    ws.cell(row=66, column=6, value="=F44")
    ws.cell(row=66, column=7, value="=G44")

    for r in range(63, 67):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c > 1:
                cell.alignment = align_right
                if cell.value is not None and not isinstance(cell.value, str):
                    cell.number_format = '#,##0'

    # ROI Formula rows
    ws.cell(row=70, column=1, value="ROI").font = font_bold
    ws.cell(row=70, column=2, value="=(H65-H64)/H64").font = font_bold
    ws.cell(row=70, column=2).number_format = '0.00'
    ws.cell(row=70, column=2).alignment = align_right
    ws.cell(row=70, column=3, value="=(Total manfaat - Total Biaya)/Total Biaya").font = font_regular

    ws.cell(row=71, column=1, value="ROI %").font = font_bold
    ws.cell(row=71, column=2, value="=B70*100").font = font_bold
    ws.cell(row=71, column=2).number_format = '0.00'
    ws.cell(row=71, column=2).alignment = align_right
    ws.cell(row=71, column=3, value="= 224.60%").font = font_bold

    # NPV Section (Row 73-80)
    ws.cell(row=73, column=1, value="NPV (Net Present Value)").font = font_bold
    
    # NPV formula row
    ws.cell(row=76, column=1, value="Tingkat bunga diskonto = 10%").font = font_bold
    ws.cell(row=76, column=2, value="=-C50+(C66/(1+(10%^1))+(D66/(1+(10%^2))+(E66/(1+(10%^3)))+(F66/(1+(10%^4)))+(G66/(1+(10%^5)))))").font = font_bold
    ws.cell(row=76, column=2).number_format = '#,##0'
    ws.cell(row=76, column=2).alignment = align_right

    # Individual PV of proceeds:
    ws.cell(row=76, column=4, value="=C66/(1+(10%^1))")
    ws.cell(row=76, column=5, value="=D66/(1+(10%^2))")
    ws.cell(row=76, column=6, value="=E66/(1+(10%^3))")
    ws.cell(row=76, column=7, value="=F66/(1+(10%^4))")
    ws.cell(row=76, column=8, value="=G66/(1+(10%^5))")

    for col in range(4, 9):
        cell = ws.cell(row=76, column=col)
        cell.number_format = '#,##0'
        cell.alignment = align_right
        cell.font = font_regular

    # Row 77 Col 2: NPV check formula
    ws.cell(row=77, column=2, value="=-B64+(SUM(D76:H76))").font = font_bold
    ws.cell(row=77, column=2).number_format = '#,##0'
    ws.cell(row=77, column=2).alignment = align_right

    # Summary text
    ws.cell(row=79, column=1, value="Dari hasil perhitungan dapat disimpulkan bahwa jumlah keuntungan yang diterima sekarang oleh perusahaan pada tahun ke 5 apabila sistem ini diterapkan adalah :").font = font_regular
    ws.cell(row=80, column=1, value="Rp 166,024,396 atau NPV lebih besar dari 0, maka proyek tersebut layak dilaksanakan").font = font_bold

    # Link pengumpulan
    ws.cell(row=85, column=1, value="link pengumpulan tugas").font = font_regular
    ws.cell(row=85, column=2, value="https://forms.gle/LurLj9dAC6iZPyHL9").font = font_regular

    # Auto-fit column widths
    ws.column_dimensions['A'].width = 52
    for c in range(2, 11):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 15

    # Save Excel
    wb.save(new_excel_path)
    print(f"Created styled Excel file up to 5 years at: {new_excel_path}")

    # Create Markdown content
    markdown_content = """# Analisis Biaya dan Manfaat Proyek ISP Billing & Management System

Dokumen ini menyajikan analisis kelayakan investasi finansial untuk implementasi **ISP Billing & Management System** (tagihan terotomatisasi MikroTik, monitoring backbone PRTG & Telegram Bot, serta presensi kehadiran staf verifikasi wajah) **hingga Tahun ke-5**.

Analisis ini didasarkan pada proyek riil Anda dengan asumsi estimasi implementasi selama **5 tahun** dan tingkat bunga diskonto sebesar **10%**.

---

## 📊 Tabel Analisis Biaya & Manfaat (Format Lengkap 5 Tahun)

| Komponen Analisis | Tahun 0 | Tahun 1 | Tahun 2 | Tahun 3 | Tahun 4 | Tahun 5 |
| :--- | :---: | :---: | :---: | :---: | :---: | :---: |
| **BIAYA-BIAYA** | | | | | | |
| **Biaya Pengembangan Sistem** | | | | | | |
| - Biaya pengadaan | Rp 40.000.000 | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 0 |
| - Biaya persiapan operasi | Rp 4.000.000 | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 0 |
| **Biaya Proyek** | | | | | | |
| - Biaya konsultan | Rp 10.000.000 | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 0 |
| - Tahap analisis sistem | Rp 8.000.000 | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 0 |
| - Tahap desain sistem | Rp 8.000.000 | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 0 |
| - Penerapan sistem | Rp 15.000.000 | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 0 |
| - Biaya penyusutan | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 0 | Rp 300.000 |
| **Total biaya proyek** | **Rp 41.000.000** | **Rp 0** | **Rp 0** | **Rp 0** | **Rp 0** | **Rp 300.000** |
| **Total Biaya Pengembangan Sistem** | **Rp 85.000.000** | **Rp 0** | **Rp 0** | **Rp 0** | **Rp 0** | **Rp 300.000** |
| | | | | | | |
| **Biaya Operasional dan Perawatan** | | | | | | |
| - Operasional | Rp 0 | Rp 3.000.000 | Rp 3.500.000 | Rp 4.000.000 | Rp 4.500.000 | Rp 5.000.000 |
| - Perawatan | Rp 0 | Rp 2.000.000 | Rp 2.500.000 | Rp 2.500.000 | Rp 3.000.000 | Rp 3.000.000 |
| **Total Biaya Operasional dan Perawatan**| **Rp 0** | **Rp 5.000.000** | **Rp 6.000.000** | **Rp 6.500.000** | **Rp 7.500.000** | **Rp 8.000.000** |
| **TOTAL BIAYA** | **Rp 85.000.000** | **Rp 5.000.000** | **Rp 6.000.000** | **Rp 6.500.000** | **Rp 7.500.000** | **Rp 8.300.000** |
| | | | | | | |
| **MANFAAT** | | | | | | |
| **Berwujud** | | | | | | |
| - Penghematan biaya operasional pershn | Rp 0 | Rp 10.000.000 | Rp 12.000.000 | Rp 15.000.000 | Rp 18.000.000 | Rp 22.000.000 |
| - Peningkatan penjualan | Rp 0 | Rp 12.000.000 | Rp 20.000.000 | Rp 30.000.000 | Rp 45.000.000 | Rp 60.000.000 |
| - Penurunan biaya persediaan | Rp 0 | Rp 4.000.000 | Rp 5.000.000 | Rp 6.000.000 | Rp 7.000.000 | Rp 8.000.000 |
| **Tak Berwujud** | | | | | | |
| - Peningkatan pelayanan | Rp 0 | Rp 5.000.000 | Rp 7.000.000 | Rp 10.000.000 | Rp 12.000.000 | Rp 15.000.000 |
| - Peningkatan kepuasan pekerjaan | Rp 0 | Rp 3.000.000 | Rp 4.000.000 | Rp 5.000.000 | Rp 7.000.000 | Rp 9.000.000 |
| - Peningkatan pengambilan keputusan | Rp 0 | Rp 4.000.000 | Rp 5.000.000 | Rp 6.000.000 | Rp 8.000.000 | Rp 10.000.000 |
| **TOTAL MANFAAT** | **Rp 0** | **Rp 38.000.000** | **Rp 53.000.000** | **Rp 72.000.000** | **Rp 97.000.000** | **Rp 124.000.000** |
| **PROCEED (TM - TB)** | **-Rp 85.000.000**| **Rp 33.000.000** | **Rp 47.000.000** | **Rp 65.500.000** | **Rp 89.500.000** | **Rp 115.700.000** |

---

## 📈 Metrik Evaluasi Kelayakan Proyek (5 Tahun)

Berdasarkan data di atas, berikut adalah hasil analisis kelayakan investasi:

### 1. Payback Period (PP)
Mengukur seberapa cepat modal investasi awal (Tahun 0) sebesar **Rp 85.000.000** dapat kembali:
* **Tahun 1 (Proceed):** Rp 33.000.000 (Sisa investasi belum kembali: Rp 52.000.000)
* **Tahun 2 (Proceed):** Rp 47.000.000 (Sisa investasi belum kembali: Rp 5.000.000)
* **Tahun 3 (Proceed):** Rp 65.500.000
* **Perhitungan:**
  $$\text{Payback Period} = 2 + \left( \frac{\text{Rp 5.000.000}}{\text{Rp 65.500.000}} \right) \approx \mathbf{2,08\text{ Tahun}}\text{ (2 Tahun 28 Hari)}$$
* *Kesimpulan:* Investasi kembali dalam waktu **2 Tahun 28 Hari** (selama berjalannya Tahun ke-3).

### 2. Return on Investment (ROI)
Mengukur total rasio keuntungan bersih terhadap biaya investasi selama 5 tahun:
* **Total Biaya Kumulatif:** Rp 118.300.000
* **Total Manfaat Kumulatif:** Rp 384.000.000
* **Perhitungan:**
  $$\text{ROI} = \frac{\text{Total Manfaat} - \text{Total Biaya}}{\text{Total Biaya}} = \frac{\text{Rp 384.000.000} - \text{Rp 118.300.000}}{\text{Rp 118.300.000}} \approx \mathbf{224,60\%}$$
* *Kesimpulan:* Proyek menghasilkan keuntungan bersih operasional sebesar **224,60%** dari total modal yang diinvestasikan.

### 3. Net Present Value (NPV) dengan Bunga 10%
Menghitung nilai bersih keuntungan saat ini dengan memperhitungkan faktor inflasi/tingkat suku bunga diskonto sebesar **10%** ($r = 0,10$):
* **PV Proceed 1:** $\frac{\text{Rp 33.000.000}}{(1 + 0,10)^1} = \text{Rp 30.000.000}$
* **PV Proceed 2:** $\frac{\text{Rp 47.000.000}}{(1 + 0,10)^2} = \text{Rp 38.842.975}$
* **PV Proceed 3:** $\frac{\text{Rp 65.500.000}}{(1 + 0,10)^3} = \text{Rp 49.211.119}$
* **PV Proceed 4:** $\frac{\text{Rp 89.500.000}}{(1 + 0,10)^4} = \text{Rp 61.129.704}$
* **PV Proceed 5:** $\frac{\text{Rp 115.700.000}}{(1 + 0,10)^5} = \text{Rp 71.840.597}$
* **Total Present Value (PV) of Proceeds:** Rp 251.024.396
* **NPV:**
  $$\text{NPV} = \text{Total PV Proceeds} - \text{Investasi Awal}$$
  $$\text{NPV} = \text{Rp 251.024.396} - \text{Rp 85.000.000} = \mathbf{Rp 166.024.396}$$
* *Kesimpulan:* Karena **NPV > 0** (positif sebesar Rp 166.024.396), proyek **ISP Billing & Management System** ini dinyatakan **sangat layak dilaksanakan** dan memberikan keuntungan yang signifikan pada tahun ke-5.

---

## 💡 Rincian Justifikasi Manfaat Proyek Anda

1. **Otomatisasi Penagihan & Isolir MikroTik (Tabel `customers`, `invoices`, `payments`)**:
   Dengan adanya integrasi API MikroTik, sistem secara otomatis mengisolasi (memblokir) IP pelanggan yang memiliki tagihan `unpaid` melewati `due_date`. Hal ini meniadakan kebutuhan peninjauan manual dan meminimalkan kerugian akibat tagihan macet yang biasanya dibiarkan aktif.
2. **Dashboard Manajemen Multi-Tenant (Tabel `users` & `packages`)**:
   Setiap owner/admin memiliki kontrol penuh atas daftar paket dan limitasi pelanggan. Data terpusat mempercepat monitoring pendapatan bulanan.
3. **Pemantauan Infrastruktur Jaringan (Tabel `backbone_devices` & PRTG)**:
   Integrasi PRTG dan bot notifikasi Telegram memungkinkan notifikasi downtime perangkat backbone dikirim secara real-time. Downtime dapat diatasi sebelum banyak pelanggan komplain.
4. **Sistem Tiket Keluhan Pelanggan Terlacak (Tabel `tickets`)**:
   Keluhan pelanggan tercatat dengan foto kendala, teknisi yang ditugaskan (`assigned_to`), status (`open`/`resolved`/`closed`), serta bukti foto perbaikan. Ini meningkatkan transparansi dan kualitas pelayanan.
   
*Catatan: Dokumen interaktif Excel dengan formula otomatis tersedia di file proyek Anda: [Analisis biaya dan manfaat.xlsx](file:///d:/xampp/htdocs/isp-billing/Analisis%20biaya%20dan%20manfaat.xlsx).*
"""

    with open(markdown_path, "w", encoding="utf-8") as f:
        f.write(markdown_content)
    print(f"Successfully generated 5 years Markdown report at: {markdown_path}")

if __name__ == "__main__":
    main()
