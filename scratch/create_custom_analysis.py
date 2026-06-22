import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    # File Paths
    workspace_dir = os.path.abspath(".")
    old_excel_path = os.path.join(workspace_dir, "Analisis biaya dan manfaat.xlsx")
    backup_excel_path = os.path.join(workspace_dir, "Analisis biaya dan manfaat (Contoh).xlsx")
    new_excel_path = os.path.join(workspace_dir, "Analisis biaya dan manfaat.xlsx")
    markdown_path = os.path.join(workspace_dir, "Analisis Biaya dan Manfaat Proyek.md")

    # 1. Back up the original file if it hasn't been backed up yet
    if os.path.exists(old_excel_path) and not os.path.exists(backup_excel_path):
        os.rename(old_excel_path, backup_excel_path)
        print(f"Backed up original example file to: {backup_excel_path}")
    elif os.path.exists(old_excel_path) and os.path.exists(backup_excel_path):
        # Delete old file to prepare for writing a fresh new styled one
        os.remove(old_excel_path)
        print(f"Removed temporary {old_excel_path} to create a fresh one.")

    # 2. Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis Biaya dan Manfaat"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="555555")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=11, bold=True, color="000000")
    font_regular = Font(name="Segoe UI", size=11, color="000000")
    font_italic = Font(name="Segoe UI", size=10, italic=True, color="333333")

    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_section = PatternFill(start_color="E6EEF8", end_color="E6EEF8", fill_type="solid")
    fill_summary = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    double_bottom_border = Border(
        top=Side(style='thin', color='1B365D'),
        bottom=Side(style='double', color='1B365D')
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ANALISIS BIAYA DAN MANFAAT PROYEK ISP BILLING & MANAGEMENT SYSTEM"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left

    ws.merge_cells("A2:E2")
    ws["A2"] = "Sistem Tagihan & Monitoring Terintegrasi MikroTik, PRTG, Telegram Bot, dan Presensi Wajah"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_left

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["Keterangan / Komponen Proyek", "Tahun 0 (Investasi)", "Tahun 1", "Tahun 2", "Tahun 3"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 1 else align_left
        cell.border = thin_border
    ws.row_dimensions[4].height = 25

    # Data Rows definition
    # Structure: (Row Type, Label, Year 0, Year 1, Year 2, Year 3, Formula/Note)
    data = [
        # SECTION: BIAYA
        ("section", "A. BIAYA-BIAYA PENGEMBANGAN (INVESTASI awal)", "", "", "", ""),
        ("item", "1. Biaya Pengadaan Router & Server Hardware Integration", 15000000, 0, 0, 0),
        ("item", "2. Biaya Persiapan Operasi & Cloud VPS Hosting (Setup)", 5000000, 0, 0, 0),
        ("item", "3. Biaya Konsultan & Analisis Perancangan Database (ERD)", 8000000, 0, 0, 0),
        ("item", "4. Biaya Software Development (Laravel, Webhook, API Integration)", 35000000, 0, 0, 0),
        ("item", "5. Biaya Sosialisasi & Pelatihan Staf/Teknisi", 2000000, 0, 0, 0),
        ("total_sec", "Total Biaya Pengembangan (Investasi Awal)", "=SUM(B6:B10)", "=SUM(C6:C10)", "=SUM(D6:D10)", "=SUM(E6:E10)"),
        
        ("empty", "", "", "", "", ""),
        
        ("section", "B. BIAYA OPERASIONAL & PEMELIHARAAN (RUNNING COSTS)", "", "", "", ""),
        ("item", "1. Sewa Cloud VPS & Domain Server", 0, 3600000, 4000000, 4400000),
        ("item", "2. Biaya API Integrasi & Telegram Bot Alert", 0, 1200000, 1200000, 1200000),
        ("item", "3. Pemeliharaan Aplikasi & Database (Bug fixing, Backups)", 0, 6000000, 6500000, 7000000),
        ("total_sec", "Total Biaya Operasional & Pemeliharaan", "=SUM(B15:B17)", "=SUM(C15:C17)", "=SUM(D15:D17)", "=SUM(E15:E17)"),
        
        ("total_all", "TOTAL BIAYA KESELURUHAN (A + B)", "=B11+B18", "=C11+C18", "=D11+D18", "=E11+E18"),
        
        ("empty", "", "", "", "", ""),
        
        # SECTION: MANFAAT
        ("section", "C. MANFAAT BERWUJUD (TANGIBLE BENEFITS)", "", "", "", ""),
        ("item", "1. Pencegahan Kebocoran Billing (Isolir Otomatis MikroTik)", 0, 18000000, 28000000, 42000000),
        ("item", "2. Efisiensi Tenaga Admin & Kolektor Tagihan", 0, 15000000, 20000000, 25000000),
        ("item", "3. Penghematan Operasional Monitoring (PRTG & Telegram Alerts)", 0, 12000000, 15000000, 18000000),
        ("total_sec", "Total Manfaat Berwujud", "=SUM(B23:B25)", "=SUM(C23:C25)", "=SUM(D23:D25)", "=SUM(E23:E25)"),
        
        ("empty", "", "", "", "", ""),
        
        ("section", "D. MANFAAT TAK BERWUJUD (INTANGIBLE BENEFITS)", "", "", "", ""),
        ("item", "1. Peningkatan Disiplin Staf (Presensi Foto Wajah & Lembur)", 0, 6000000, 9000000, 12000000),
        ("item", "2. Loyalitas & Kepuasan Pelanggan (Sistem Tiket Cepat + Foto)", 0, 10000000, 16000000, 22000000),
        ("item", "3. Peningkatan Akurasi Keputusan (Dashboard Multi-tenant)", 0, 8000000, 12000000, 16000000),
        ("total_sec", "Total Manfaat Tak Berwujud", "=SUM(B30:B32)", "=SUM(C30:C32)", "=SUM(D30:D32)", "=SUM(E30:E32)"),
        
        ("total_all", "TOTAL MANFAAT KESELURUHAN (C + D)", "=B26+B33", "=C26+C33", "=D26+D33", "=E26+E33"),
        
        ("empty", "", "", "", "", ""),
        
        # PROCEED
        ("total_all", "PROCEED / SELISIH MANFAAT & BIAYA (TM - TB)", "=B34-B19", "=C34-C19", "=D34-D19", "=E34-E19"),
    ]

    current_row = 5
    for row_data in data:
        row_type, label, y0, y1, y2, y3 = row_data
        
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=y0)
        ws.cell(row=current_row, column=3, value=y1)
        ws.cell(row=current_row, column=4, value=y2)
        ws.cell(row=current_row, column=5, value=y3)

        # Style according to row type
        if row_type == "section":
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            cell = ws.cell(row=current_row, column=1)
            cell.font = font_bold
            cell.fill = fill_section
            cell.alignment = align_left
            ws.row_dimensions[current_row].height = 22
        elif row_type == "item":
            ws.cell(row=current_row, column=1).font = font_regular
            ws.cell(row=current_row, column=1).alignment = align_left
            for col in range(2, 6):
                c = ws.cell(row=current_row, column=col)
                c.font = font_regular
                c.alignment = align_right
                c.number_format = '#,##0'
                c.border = thin_border
            ws.row_dimensions[current_row].height = 20
        elif row_type == "total_sec":
            ws.cell(row=current_row, column=1).font = font_bold
            ws.cell(row=current_row, column=1).alignment = align_left
            for col in range(2, 6):
                c = ws.cell(row=current_row, column=col)
                c.font = font_bold
                c.alignment = align_right
                c.number_format = '#,##0'
                c.border = thin_border
            ws.row_dimensions[current_row].height = 20
        elif row_type == "total_all":
            ws.cell(row=current_row, column=1).font = font_bold
            ws.cell(row=current_row, column=1).alignment = align_left
            ws.cell(row=current_row, column=1).fill = fill_summary
            for col in range(2, 6):
                c = ws.cell(row=current_row, column=col)
                c.font = font_bold
                c.alignment = align_right
                c.number_format = '#,##0'
                c.fill = fill_summary
                c.border = double_bottom_border
            ws.row_dimensions[current_row].height = 22
        elif row_type == "empty":
            ws.row_dimensions[current_row].height = 10
            
        current_row += 1

    # Analysis calculations section in Excel
    current_row += 1
    ws.cell(row=current_row, column=1, value="E. ANALISIS KELAYAKAN FINANSIAL").font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    ws.cell(row=current_row, column=1).fill = fill_section
    ws.row_dimensions[current_row].height = 25
    current_row += 1

    # Discount Rate Row
    ws.cell(row=current_row, column=1, value="1. Tingkat Bunga Diskonto (r)").font = font_bold
    ws.cell(row=current_row, column=2, value=0.10)
    ws.cell(row=current_row, column=2).number_format = '0%'
    ws.cell(row=current_row, column=2).font = font_bold
    ws.cell(row=current_row, column=2).alignment = align_right
    current_row += 1

    # NPV Row
    # NPV formula in excel: =NPV(rate, Year1_Proceed, Year2_Proceed, Year3_Proceed) + Year0_Proceed
    # Row indices for proceeds: Year0=B36, Year1=C36, Year2=D36, Year3=E36
    # Note: B36 is negative (investasi awal)
    ws.cell(row=current_row, column=1, value="2. Net Present Value (NPV)").font = font_bold
    # NPV formula referencing C36, D36, E36 and adding B36 (which is negative)
    # The rate is in B39
    ws.cell(row=current_row, column=2, value="=NPV(B39, C36, D36, E36) + B36")
    ws.cell(row=current_row, column=2).number_format = '#,##0'
    ws.cell(row=current_row, column=2).font = font_bold
    ws.cell(row=current_row, column=2).alignment = align_right
    ws.cell(row=current_row, column=3, value="Proyek Layak dilaksanakan jika NPV > 0").font = font_italic
    current_row += 1

    # ROI Row
    # ROI = (Total Manfaat 3 thn - Total Biaya 3 thn) / Total Biaya 3 thn
    # Total Biaya: =SUM(B19:E19)
    # Total Manfaat: =SUM(B34:E34)
    ws.cell(row=current_row, column=1, value="3. Return on Investment (ROI)").font = font_bold
    ws.cell(row=current_row, column=2, value="=(SUM(B34:E34)-SUM(B19:E19))/SUM(B19:E19)")
    ws.cell(row=current_row, column=2).number_format = '0.00%'
    ws.cell(row=current_row, column=2).font = font_bold
    ws.cell(row=current_row, column=2).alignment = align_right
    current_row += 1

    # Payback Period Row
    ws.cell(row=current_row, column=1, value="4. Payback Period (Tahun)").font = font_bold
    # Proceed thn 1 is C36 (58.200.000). Investasi is -B36 (65.000.000).
    # Sisa = 65M - 58.2M = 6.8M. Proceed thn 2 is D36 (85.300.000).
    # PP = 1 + (6.8M / 85.3M) = 1.08 tahun. We'll write the formula or the value for simplicity.
    ws.cell(row=current_row, column=2, value=1.08)
    ws.cell(row=current_row, column=2).number_format = '0.00'
    ws.cell(row=current_row, column=2).font = font_bold
    ws.cell(row=current_row, column=2).alignment = align_right
    ws.cell(row=current_row, column=3, value="1 Tahun 29 Hari").font = font_italic
    current_row += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid using length of formula strings for width
            val = str(cell.value or '')
            if val.startswith('='):
                val = "Rp 999,999,999" # Placeholder size for formulas
            if cell.number_format and ('#,##0' in cell.number_format or '0%' in cell.number_format):
                val = "Rp 999,999,999"
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.column_dimensions['A'].width = 55

    # Save Excel
    wb.save(new_excel_path)
    print(f"Created styled Excel file at: {new_excel_path}")

    # 3. Create Markdown File
    markdown_content = """# Analisis Biaya dan Manfaat Proyek ISP Billing & Management System

Dokumen ini menyajikan analisis kelayakan investasi finansial untuk implementasi **ISP Billing & Management System** (tagihan terotomatisasi MikroTik, monitoring backbone PRTG & Telegram Bot, serta presensi kehadiran staf verifikasi wajah).

Analisis ini didasarkan pada proyek riil Anda dengan asumsi estimasi implementasi selama **3 tahun** dan tingkat bunga diskonto sebesar **10%**.

---

## 📊 Tabel Analisis Biaya & Manfaat

| Komponen Analisis | Tahun 0 (Investasi) | Tahun 1 | Tahun 2 | Tahun 3 |
| :--- | :---: | :---: | :---: | :---: |
| **A. BIAYA PENGEMBANGAN (INVESTASI)** | | | | |
| 1. Biaya Pengadaan Router & Server Hardware | Rp 15.000.000 | Rp 0 | Rp 0 | Rp 0 |
| 2. Biaya Persiapan Operasi & Cloud VPS (Setup) | Rp 5.000.000 | Rp 0 | Rp 0 | Rp 0 |
| 3. Biaya Konsultan & Analisis Perancangan Database (ERD) | Rp 8.000.000 | Rp 0 | Rp 0 | Rp 0 |
| 4. Biaya Software Development (Laravel, API, Bot) | Rp 35.000.000 | Rp 0 | Rp 0 | Rp 0 |
| 5. Biaya Sosialisasi & Pelatihan Staf/Teknisi | Rp 2.000.000 | Rp 0 | Rp 0 | Rp 0 |
| **Total Biaya Pengembangan (Investasi Awal)** | **Rp 65.000.000** | **Rp 0** | **Rp 0** | **Rp 0** |
| | | | | |
| **B. BIAYA OPERASIONAL & MAINTENANCE** | | | | |
| 1. Sewa Cloud VPS & Domain Server | Rp 0 | Rp 3.600.000 | Rp 4.000.000 | Rp 4.400.000 |
| 2. Biaya API Integrasi & Telegram Bot Alert | Rp 0 | Rp 1.200.000 | Rp 1.200.000 | Rp 1.200.000 |
| 3. Pemeliharaan Aplikasi & Database (Bug fixes) | Rp 0 | Rp 6.000.000 | Rp 6.500.000 | Rp 7.000.000 |
| **Total Biaya Operasional & Pemeliharaan** | **Rp 0** | **Rp 10.800.000** | **Rp 11.700.000** | **Rp 12.600.000** |
| **TOTAL BIAYA KESELURUHAN (A + B)** | **Rp 65.000.000** | **Rp 10.800.000** | **Rp 11.700.000** | **Rp 12.600.000** |
| | | | | |
| **C. MANFAAT BERWUJUD (TANGIBLE)** | | | | |
| 1. Pencegahan Kebocoran Billing (Isolir Otomatis MikroTik)| Rp 0 | Rp 18.000.000 | Rp 28.000.000 | Rp 42.000.000 |
| 2. Efisiensi Tenaga Admin & Kolektor Tagihan | Rp 0 | Rp 15.000.000 | Rp 20.000.000 | Rp 25.000.000 |
| 3. Penghematan Monitoring (PRTG & Telegram Alerts) | Rp 0 | Rp 12.000.000 | Rp 15.000.000 | Rp 18.000.000 |
| **Total Manfaat Berwujud** | **Rp 0** | **Rp 45.000.000** | **Rp 63.000.000** | **Rp 85.000.000** |
| | | | | |
| **D. MANFAAT TAK BERWUJUD (INTANGIBLE)** | | | | |
| 1. Peningkatan Disiplin Staf (Presensi Foto Wajah) | Rp 0 | Rp 6.000.000 | Rp 9.000.000 | Rp 12.000.000 |
| 2. Loyalitas & Kepuasan Pelanggan (Sistem Tiket & Foto) | Rp 0 | Rp 10.000.000 | Rp 16.000.000 | Rp 22.000.000 |
| 3. Akurasi Keputusan (Dashboard Multi-tenant) | Rp 0 | Rp 8.000.000 | Rp 12.000.000 | Rp 16.000.000 |
| **Total Manfaat Tak Berwujud** | **Rp 0** | **Rp 24.000.000** | **Rp 37.000.000** | **Rp 50.000.000** |
| **TOTAL MANFAAT KESELURUHAN (C + D)** | **Rp 0** | **Rp 69.000.000** | **Rp 100.000.000**| **Rp 135.000.000**|
| | | | | |
| **PROCEED (SELISIH MANFAAT & BIAYA)** | **-Rp 65.000.000**| **Rp 58.200.000** | **Rp 88.300.000** | **Rp 122.400.000**|

---

## 📈 Metrik Evaluasi Kelayakan Proyek

Berdasarkan data di atas, berikut adalah hasil analisis kelayakan investasi:

### 1. Payback Period (PP)
Mengukur seberapa cepat modal investasi awal (Tahun 0) sebesar **Rp 65.000.000** dapat kembali:
* **Tahun 1 (Proceed):** Rp 58.200.000 (Sisa investasi belum kembali: Rp 6.800.000)
* **Tahun 2 (Proceed):** Rp 88.300.000
* **Perhitungan:**
  $$\text{Payback Period} = 1 + \left( \frac{\text{Rp 6.800.000}}{\text{Rp 88.300.000}} \right) \approx \mathbf{1,08\text{ Tahun}}\text{ (1 Tahun 29 Hari)}$$
* *Kesimpulan:* Investasi kembali dalam waktu **sangat singkat** (hanya dalam waktu 1 tahun 29 hari) karena efisiensi penagihan yang tinggi.

### 2. Return on Investment (ROI)
Mengukur total rasio keuntungan bersih terhadap biaya investasi selama 3 tahun:
* **Total Biaya Kumulatif:** Rp 100.100.000
* **Total Manfaat Kumulatif:** Rp 304.000.000
* **Perhitungan:**
  $$\text{ROI} = \frac{\text{Total Manfaat} - \text{Total Biaya}}{\text{Total Biaya}} = \frac{\text{Rp 304.000.000} - \text{Rp 100.100.000}}{\text{Rp 100.100.000}} \approx \mathbf{203,70\%}$$
* *Kesimpulan:* Proyek menghasilkan keuntungan operasional sebesar **203,70%** dari total modal yang dikeluarkan.

### 3. Net Present Value (NPV)
Menghitung nilai bersih keuntungan saat ini dengan memperhitungkan faktor inflasi/tingkat suku bunga diskonto sebesar **10%** ($r = 0,10$):
* **PV Proceed 1:** $\frac{\text{Rp 58.200.000}}{(1 + 0,10)^1} \approx \text{Rp 52.909.091}$
* **PV Proceed 2:** $\frac{\text{Rp 88.300.000}}{(1 + 0,10)^2} \approx \text{Rp 72.975.207}$
* **PV Proceed 3:** $\frac{\text{Rp 122.400.000}}{(1 + 0,10)^3} \approx \text{Rp 91.960.932}$
* **Total Present Value (PV):** Rp 217.845.230
* **NPV:**
  $$\text{NPV} = \text{Total PV Proceeds} - \text{Investasi Awal}$$
  $$\text{NPV} = \text{Rp 217.845.230} - \text{Rp 65.000.000} = \mathbf{Rp 152.845.230}$$
* *Kesimpulan:* Karena **NPV > 0** (bernilai positif sebesar Rp 152.845.230), maka proyek pengembangan **ISP Billing & Management System** ini **sangat layak dan menguntungkan** untuk dilaksanakan.

---

## 💡 Rincian Justifikasi Manfaat Proyek Anda

1. **Otomatisasi Penagihan & Isolir MikroTik (Tabel `customers`, `invoices`, `payments`)**:
   Dengan adanya integrasi API MikroTik, sistem secara otomatis mengisolasi (memblokir) IP pelanggan yang memiliki tagihan `unpaid` melewati `due_date`. Hal ini meniadakan kebutuhan peninjauan manual dan meminimalkan kerugian akibat tagihan macet yang biasanya dibiarkan aktif.
2. **Dashboard Manajemen Multi-Tenant (Tabel `users` & `packages`)**:
   Setiap owner/admin memiliki kontrol penuh atas daftar paket dan limitasi pelanggan. Data terpusat mempercepat monitoring pendapatan bulanan.
3. **Pemantauan Infrastruktur Jaringan (Tabel `backbone_devices` & PRTG)**:
   Integrasi PRTG dan bot notifikasi Telegram memungkinkan notifikasi downtime perangkat backbone dikirim secara real-time. Downtime dapat diatasi sebelum banyak pelanggan komplain.
4. **Sistem Tiket Keluhan Pelanggan Terlacak (Tabel `tickets`)**:
   Keluhan pelanggan tercatat dengan foto kendala, teknisi yang ditugaskan (`assigned_to`), status (`open`/`resolved`/`closed`), serta bukti foto perbaikan. Ini meningkatkan transparansi dan kualitas pelayanan.
   
*Catatan: Dokumen interaktif Excel dengan formula otomatis tersedia di file proyek Anda: [Analisis biaya dan manfaat.xlsx](file:///d:/xampp/htdocs/isp-billing/Analisis%20biaya%20dan%20manfaat.xlsx).*
"""

    with open(markdown_path, "w", encoding="utf-8") as f:
        f.write(markdown_content)
    print(f"Created Markdown description file at: {markdown_path}")

if __name__ == "__main__":
    main()
